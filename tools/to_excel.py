import os
import json
import ast
from core.app_config import settings
from openpyxl import Workbook, load_workbook
from ._common import parse_input, resolve_path

DESCRIPTION = "Writes data to an .xlsx file. Can create a new sheet from a list of dictionaries (rows) or from plain text."
ARGS_SCHEMA = '''{
  "output_path": "<string: path for the new .xlsx file>",
  "rows": "[<object>: optional, a list of data rows, e.g., [{'col1': 'valA'}, {'col1': 'valB'}]]",
  "text": "<string: optional, plain text to write line-by-line into the first column>",
  "template_path": "<string: optional, path to an existing .xlsx file to use as a template for headers>"
}'''

def run(action_input):
    if not settings.get("ALLOW_FILE_CREATE", True):
        return "Error: File creation is disabled by permissions."

    try:
        args = parse_input(action_input)
        output_path = resolve_path(args.get("output_path", ""))
        rows_data = args.get("rows")
        text_data = args.get("text")
        template_path = resolve_path(args.get("template_path", ""))

        if not output_path:
            return "Error: 'output_path' is a required argument."
        if not rows_data and not text_data:
            return "Error: You must provide either 'rows' or 'text' data to write."

        # --- Build Workbook ---
        if template_path and os.path.exists(template_path):
            wb = load_workbook(template_path)
            ws = wb.active
            start_row = ws.max_row + 1
            headers = [c.value for c in ws[1]]
        else:
            wb = Workbook()
            ws = wb.active
            start_row = 1
            headers = []

        # --- Write Data ---
        if rows_data and isinstance(rows_data, list):
            if not headers: # Create headers from data if no template
                headers = sorted({k for row in rows_data for k in row})
                for c_idx, header_val in enumerate(headers, 1):
                    ws.cell(row=1, column=c_idx, value=header_val)
                start_row = 2
            
            for r_idx, row_dict in enumerate(rows_data, start=start_row):
                for c_idx, header_val in enumerate(headers, 1):
                    ws.cell(row=r_idx, column=c_idx, value=row_dict.get(header_val))

        elif text_data and isinstance(text_data, str):
            for i, line in enumerate(text_data.strip().splitlines(), start=start_row):
                ws.cell(row=i, column=1, value=line)
        
        # --- Save ---
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        return f"Successfully saved spreadsheet to {output_path}"

    except Exception as e:
        return f"Error writing to Excel file: {e}"
